"""
报告格式化器

将报告转换为不同格式
"""

import json
from abc import ABC, abstractmethod
from typing import Any
import logging

from .report_generator import AnalysisReport

logger = logging.getLogger("mingjing.formatters")


class BaseFormatter(ABC):
    """格式化器基类"""
    
    @abstractmethod
    def format(self, report: AnalysisReport, *args, **kwargs) -> Any:
        """格式化报告"""
        pass


class JsonFormatter(BaseFormatter):
    """JSON格式化器"""
    
    def format(self, report: AnalysisReport, indent: int = 2) -> str:
        """格式化为JSON字符串"""
        return report.to_json(indent=indent)


class HtmlFormatter(BaseFormatter):
    """HTML格式化器"""
    
    TEMPLATE = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body class="bg-gray-50 min-h-screen">
    <div class="max-w-7xl mx-auto px-4 py-8">
        <!-- 头部 -->
        <header class="mb-8">
            <h1 class="text-3xl font-bold text-gray-900">{title}</h1>
            <p class="text-gray-600 mt-2">{description}</p>
            <div class="flex items-center gap-4 mt-4 text-sm text-gray-500">
                <span>报告ID: {report_id}</span>
                <span>生成时间: {created_at}</span>
            </div>
        </header>

        <!-- 风险概览 -->
        <section class="mb-8">
            <h2 class="text-xl font-semibold text-gray-800 mb-4">风险概览</h2>
            <div class="grid grid-cols-1 md:grid-cols-4 gap-4">
                <div class="bg-white rounded-lg shadow p-6">
                    <div class="text-sm text-gray-500">风险等级</div>
                    <div class="text-2xl font-bold {risk_color}">{risk_level_cn}</div>
                </div>
                <div class="bg-white rounded-lg shadow p-6">
                    <div class="text-sm text-gray-500">风险分数</div>
                    <div class="text-2xl font-bold text-gray-900">{risk_score}/100</div>
                </div>
                <div class="bg-white rounded-lg shadow p-6">
                    <div class="text-sm text-gray-500">识别实体数</div>
                    <div class="text-2xl font-bold text-gray-900">{total_entities}</div>
                </div>
                <div class="bg-white rounded-lg shadow p-6">
                    <div class="text-sm text-gray-500">处理文件数</div>
                    <div class="text-2xl font-bold text-gray-900">{processed_files}/{total_files}</div>
                </div>
            </div>
        </section>

        <!-- 统计图表 -->
        <section class="mb-8">
            <h2 class="text-xl font-semibold text-gray-800 mb-4">统计分析</h2>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                <div class="bg-white rounded-lg shadow p-6">
                    <h3 class="text-lg font-medium text-gray-700 mb-4">实体类型分布</h3>
                    <canvas id="entityChart" height="200"></canvas>
                </div>
                <div class="bg-white rounded-lg shadow p-6">
                    <h3 class="text-lg font-medium text-gray-700 mb-4">文件类型分布</h3>
                    <canvas id="fileChart" height="200"></canvas>
                </div>
            </div>
        </section>

        <!-- 风险摘要 -->
        <section class="mb-8">
            <h2 class="text-xl font-semibold text-gray-800 mb-4">风险摘要</h2>
            <div class="space-y-3">
                {risk_summaries_html}
            </div>
        </section>

        <!-- 详细结果 -->
        <section class="mb-8">
            <h2 class="text-xl font-semibold text-gray-800 mb-4">详细结果</h2>
            <div class="bg-white rounded-lg shadow overflow-hidden">
                <table class="min-w-full divide-y divide-gray-200">
                    <thead class="bg-gray-50">
                        <tr>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase">文件</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase">类型</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase">敏感信息</th>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase">置信度</th>
                        </tr>
                    </thead>
                    <tbody class="bg-white divide-y divide-gray-200">
                        {results_table_html}
                    </tbody>
                </table>
            </div>
        </section>
    </div>

    <script>
        // 实体类型图表
        const entityData = {entity_chart_data};
        new Chart(document.getElementById('entityChart'), {{
            type: 'doughnut',
            data: {{
                labels: entityData.labels,
                datasets: [{{
                    data: entityData.values,
                    backgroundColor: [
                        '#ef4444', '#f97316', '#eab308', '#22c55e', '#06b6d4',
                        '#3b82f6', '#8b5cf6', '#ec4899', '#6b7280', '#14b8a6'
                    ]
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ position: 'right' }}
                }}
            }}
        }});

        // 文件类型图表
        const fileData = {file_chart_data};
        new Chart(document.getElementById('fileChart'), {{
            type: 'bar',
            data: {{
                labels: fileData.labels,
                datasets: [{{
                    label: '文件数量',
                    data: fileData.values,
                    backgroundColor: '#3b82f6'
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ display: false }}
                }}
            }}
        }});
    </script>
</body>
</html>'''
    
    RISK_COLORS = {
        "critical": "text-red-600",
        "high": "text-orange-600",
        "medium": "text-yellow-600",
        "low": "text-green-600",
    }
    
    RISK_LEVEL_CN = {
        "critical": "严重",
        "high": "高",
        "medium": "中",
        "low": "低",
    }
    
    def format(self, report: AnalysisReport) -> str:
        """格式化为HTML"""
        # 风险摘要HTML
        risk_summaries_html = ""
        for summary in report.risk_summaries[:10]:  # 最多显示10条
            level_color = self.RISK_COLORS.get(summary.level, "text-gray-600")
            risk_summaries_html += f'''
                <div class="bg-white rounded-lg shadow p-4 flex items-center justify-between">
                    <div>
                        <span class="{level_color} font-medium">[{self.RISK_LEVEL_CN.get(summary.level, summary.level)}]</span>
                        <span class="text-gray-700 ml-2">{summary.description}</span>
                    </div>
                    <span class="text-sm text-gray-500">{len(summary.affected_files)} 个文件</span>
                </div>
            '''
        
        # 结果表格HTML
        results_table_html = ""
        for fr in report.file_results:
            for entity in fr.entities[:100]:  # 每个文件最多显示100条
                masked_text = entity.text[:20] + "..." if len(entity.text) > 20 else entity.text
                score_color = "text-green-600" if entity.score >= 0.8 else "text-yellow-600" if entity.score >= 0.5 else "text-red-600"
                results_table_html += f'''
                    <tr>
                        <td class="px-6 py-4 whitespace-nowrap text-sm text-gray-900">{fr.filename}</td>
                        <td class="px-6 py-4 whitespace-nowrap">
                            <span class="px-2 py-1 text-xs font-medium rounded bg-blue-100 text-blue-800">{entity.entity_type}</span>
                        </td>
                        <td class="px-6 py-4 text-sm text-gray-500 font-mono">{masked_text}</td>
                        <td class="px-6 py-4 whitespace-nowrap text-sm {score_color}">{entity.score:.0%}</td>
                    </tr>
                '''
        
        # 图表数据
        entity_chart_data = json.dumps({
            "labels": list(report.statistics.entity_type_counts.keys()),
            "values": list(report.statistics.entity_type_counts.values()),
        })
        
        file_chart_data = json.dumps({
            "labels": list(report.statistics.file_type_counts.keys()),
            "values": list(report.statistics.file_type_counts.values()),
        })
        
        # 填充模板
        html = self.TEMPLATE.format(
            title=report.title,
            description=report.description,
            report_id=report.report_id,
            created_at=report.created_at,
            risk_level_cn=self.RISK_LEVEL_CN.get(report.risk_level, report.risk_level),
            risk_color=self.RISK_COLORS.get(report.risk_level, "text-gray-600"),
            risk_score=report.risk_score,
            total_entities=report.statistics.total_entities,
            processed_files=report.statistics.processed_files,
            total_files=report.statistics.total_files,
            risk_summaries_html=risk_summaries_html,
            results_table_html=results_table_html,
            entity_chart_data=entity_chart_data,
            file_chart_data=file_chart_data,
        )
        
        return html


class ExcelFormatter(BaseFormatter):
    """Excel格式化器"""
    
    def format(self, report: AnalysisReport, filepath: str) -> None:
        """格式化为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl未安装，无法导出Excel。请运行: pip install openpyxl")
            return
        
        wb = openpyxl.Workbook()
        
        # 概览Sheet
        ws_overview = wb.active
        ws_overview.title = "概览"
        self._write_overview(ws_overview, report)
        
        # 详细结果Sheet
        ws_details = wb.create_sheet("详细结果")
        self._write_details(ws_details, report)
        
        # 统计Sheet
        ws_stats = wb.create_sheet("统计")
        self._write_statistics(ws_stats, report)
        
        wb.save(filepath)
    
    def _write_overview(self, ws, report: AnalysisReport) -> None:
        """写入概览"""
        from openpyxl.styles import Font, PatternFill
        
        header_font = Font(bold=True, size=14)
        
        ws['A1'] = report.title
        ws['A1'].font = header_font
        
        ws['A3'] = "报告ID"
        ws['B3'] = report.report_id
        ws['A4'] = "生成时间"
        ws['B4'] = report.created_at
        ws['A5'] = "风险等级"
        ws['B5'] = report.risk_level
        ws['A6'] = "风险分数"
        ws['B6'] = report.risk_score
        ws['A7'] = "总实体数"
        ws['B7'] = report.statistics.total_entities
        ws['A8'] = "处理文件数"
        ws['B8'] = report.statistics.processed_files
        ws['A9'] = "失败文件数"
        ws['B9'] = report.statistics.failed_files
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
    
    def _write_details(self, ws, report: AnalysisReport) -> None:
        """写入详细结果"""
        from openpyxl.styles import Font
        
        # 表头
        headers = ["文件名", "实体类型", "敏感信息", "起始位置", "结束位置", "置信度", "已验证", "来源"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # 数据
        row = 2
        for fr in report.file_results:
            for entity in fr.entities:
                ws.cell(row=row, column=1, value=fr.filename)
                ws.cell(row=row, column=2, value=entity.entity_type)
                ws.cell(row=row, column=3, value=entity.text)
                ws.cell(row=row, column=4, value=entity.start)
                ws.cell(row=row, column=5, value=entity.end)
                ws.cell(row=row, column=6, value=f"{entity.score:.2%}")
                ws.cell(row=row, column=7, value="是" if entity.verified else "否")
                ws.cell(row=row, column=8, value=entity.source)
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['H'].width = 30
    
    def _write_statistics(self, ws, report: AnalysisReport) -> None:
        """写入统计信息"""
        from openpyxl.styles import Font
        
        ws['A1'] = "实体类型统计"
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A3'] = "实体类型"
        ws['B3'] = "数量"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        row = 4
        for entity_type, count in report.statistics.entity_type_counts.items():
            ws.cell(row=row, column=1, value=entity_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        row += 2
        ws.cell(row=row, column=1, value="文件类型统计")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        row += 2
        ws.cell(row=row, column=1, value="文件类型")
        ws.cell(row=row, column=2, value="数量")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        row += 1
        for file_type, count in report.statistics.file_type_counts.items():
            ws.cell(row=row, column=1, value=file_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
